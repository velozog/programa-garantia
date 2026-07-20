import re
import math
from termcolor import colored
import colorama
from typing import Literal
import json
import copy
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import os
from rich.table import Table
from rich.console import Console
from datetime import datetime, timedelta
import requests
from random import randint
from rich.console import Console
from rich.panel import Panel
from rich.align import Align

try:
    from prompt_toolkit import prompt
    from prompt_toolkit.styles import Style
    from prompt_toolkit.history import InMemoryHistory
    from prompt_toolkit.validation import Validator, ValidationError
    PROMPT_TOOLKIT_AVAILABLE = True
except ImportError:
    prompt = None
    Style = None
    InMemoryHistory = None
    Validator = object
    ValidationError = type("ValidationError", (Exception,), {})
    PROMPT_TOOLKIT_AVAILABLE = False


DYID_REGEX_LETTER = re.compile(r"^DyG\.(\d{2})\.([a-zA-Z]{1}\d{2})\.(\d{4})$")
DYID_REGEX = re.compile(r"^DyG\.(\d{2})\.(\d{3})\.(\d{4})$")
DYID_REGEX_LEGACY = re.compile(r"^DyG\.(\d{3})\.(\d{4})$")
DYID_REGEX_V2 = re.compile(r"^DyG(\d)\.(\d{2})\.([A-Z])(\d{4})$")

colorama.init(autoreset=True)

SCREEN_SIZE = 80

COLORS = {
    "BLACK": colorama.Back.BLACK,
    "RED": colorama.Back.RED,
    "GREEN": colorama.Back.GREEN,
    "YELLOW": colorama.Back.YELLOW,
    "BLUE": colorama.Back.BLUE,
    "MAGENTA": colorama.Back.MAGENTA,
    "CYAN": colorama.Back.CYAN,
    "WHITE": colorama.Back.WHITE,
    "RESET": colorama.Back.RESET,
}

Color = Literal[
    "black",
    "grey",
    "red",
    "green",
    "yellow",
    "blue",
    "magenta",
    "cyan",
    "light_grey",
    "dark_grey",
    "light_red",
    "light_green",
    "light_yellow",
    "light_blue",
    "light_magenta",
    "light_cyan",
    "white",
]

def dynagatewayToolsLogo(version):
    dynaloggerTools = """

░██████╗██╗░██████╗████████╗███████╗███╗░░░███╗░█████╗░  ██████╗░███████╗
██╔════╝██║██╔════╝╚══██╔══╝██╔════╝████╗░████║██╔══██╗  ██╔══██╗██╔════╝
╚█████╗░██║╚█████╗░░░░██║░░░█████╗░░██╔████╔██║███████║  ██║░░██║█████╗░░
░╚═══██╗██║░╚═══██╗░░░██║░░░██╔══╝░░██║╚██╔╝██║██╔══██║  ██║░░██║██╔══╝░░
██████╔╝██║██████╔╝░░░██║░░░███████╗██║░╚═╝░██║██║░░██║  ██████╔╝███████╗
╚═════╝░╚═╝╚═════╝░░░░╚═╝░░░╚══════╝╚═╝░░░░░╚═╝╚═╝░░╚═╝  ╚═════╝░╚══════╝

░█████╗░░██████╗░██████╗██╗░██████╗████████╗███████╗███╗░░██╗░█████╗░██╗░█████╗░
██╔══██╗██╔════╝██╔════╝██║██╔════╝╚══██╔══╝██╔════╝████╗░██║██╔══██╗██║██╔══██╗
███████║╚█████╗░╚█████╗░██║╚█████╗░░░░██║░░░█████╗░░██╔██╗██║██║░░╚═╝██║███████║
██╔══██║░╚═══██╗░╚═══██╗██║░╚═══██╗░░░██║░░░██╔══╝░░██║╚████║██║░░██╗██║██╔══██║
██║░░██║██████╔╝██████╔╝██║██████╔╝░░░██║░░░███████╗██║░╚███║╚█████╔╝██║██║░░██║
╚═╝░░╚═╝╚═════╝░╚═════╝░╚═╝╚═════╝░░░░╚═╝░░░╚══════╝╚═╝░░╚══╝░╚════╝░╚═╝╚═╝░░╚═╝

████████╗███████╗░█████╗░███╗░░██╗██╗░█████╗░░█████╗░
╚══██╔══╝██╔════╝██╔══██╗████╗░██║██║██╔══██╗██╔══██╗
░░░██║░░░█████╗░░██║░░╚═╝██╔██╗██║██║██║░░╚═╝███████║
░░░██║░░░██╔══╝░░██║░░██╗██║╚████║██║██║░░██╗██╔══██║
░░░██║░░░███████╗╚█████╔╝██║░╚███║██║╚█████╔╝██║░░██║
░░░╚═╝░░░╚══════╝░╚════╝░╚═╝░░╚══╝╚═╝░╚════╝░╚═╝░░╚═╝
"""
    print(colored(dynaloggerTools, "cyan"))
    print(colored(version, "cyan"))

def printSubTitle(
    message: str = "",
    color: str = "white",
    screenSize: int = SCREEN_SIZE,
):
    RESET = "\033[0m"
    auxSize = (screenSize - len(message)) / 2.0
    auxSize = auxSize if auxSize >= 0 else 0

    textToDisplay = str().join(list("-" for i in range(math.floor(auxSize))))
    textToDisplay += message
    textToDisplay += str().join(list("-" for i in range(math.ceil(auxSize))))
    print()
    print(colored("\n" + textToDisplay, color)+RESET)
    print()




def validate_dyid(dyid: str) -> bool:
    validate = False
    if (
        DYID_REGEX.match(dyid)
        or DYID_REGEX_LEGACY.match(dyid)
        or DYID_REGEX_LETTER.match(dyid)
        or DYID_REGEX_V2.match(dyid)
    ):
        validate = True
    return validate

def generic_prompt(options: list, menu: list) -> int:
    """Creates a menu with given options and handles user input

    Args:
        options (list): Options to be selected
        menu (list): Menu to be printed

    Raises:
        ValueError: Not an option

    Returns:
        int: Valid option
    """
    while True:
        for option in menu:
            print(colored(f"{option}", "cyan"))

        try:
            resposta = perguntar("Selecione uma das opções acima:", titulo="Menu", obrigatorio=True, strip=True)
            user_choice = int(resposta)

            if user_choice not in options:
                raise ValueError

            return user_choice

        except ValueError:
            printSubTitle("Opção inválida!", "red")

def validate_info(text: str) -> bool:
    """Ask user to validate informations.

    Args:
        text (str): Question to be asked

    Returns:
        bool: Confirmation
    """
    while True:
        resposta = perguntar(text, titulo="Confirmação Use S/N ou 1/0.", obrigatorio=False, strip=True)
        resposta_normalizada = resposta.strip().lower().replace("ã", "a").replace("á", "a").replace("â", "a")

        if resposta_normalizada in {"s", "sim", "y", "yes", "1"}:
            return True
        if resposta_normalizada in {"n", "nao", "no", "0"}:
            return False

        printSubTitle("Resposta inválida! Use S/N ou 1/0.", "red")

def scanOneId():
    while True:
        dyid = perguntar("Insira o Dyid:", titulo="Entrada", obrigatorio=True, strip=True)
        if validate_dyid(dyid):
            return dyid
        else:
            printSubTitle("Formato invalido! Ignorando ID", "red")

def identificar_modelo(dyid):
    """
    Identifica o modelo do produto baseado no ID.

    Args:
        id_produto: ID do produto (ex: 'DyG.10.A18.1545')

    Returns:
        str: Nome do modelo correspondente ou None se não encontrado

    """
    while True:
        if validate_dyid(dyid):
            break
        else:
            printSubTitle("Formato invalido! Ignorando ID", "red")
            dyid = perguntar("Digite o DYID:", titulo="Entrada", obrigatorio=True, strip=True)
    
    with open("config.json", "r", encoding="utf-8") as f:
        config = json.load(f)

    modelos = config["modelos"]

    # Ordenar as chaves por tamanho (decrescente) para verificar prefixos mais longos primeiro
    chaves_ordenadas = sorted(modelos.keys(), key=len, reverse=True)

    for prefixo in chaves_ordenadas:
        if dyid.startswith(prefixo):
            return modelos[prefixo], dyid

    return None, dyid

def validate_input_min_length(mensagem, min_length=3):
    """
    Valida se o input tem o comprimento mínimo especificado.

    Args:
        mensagem: Mensagem para exibir ao usuário
        min_length: Comprimento mínimo exigido (padrão: 3)

    Returns:
        str: Input validado com comprimento mínimo
    """
    while True:
        valor = perguntar(mensagem, titulo="Entrada", obrigatorio=True, strip=True)
        if len(valor) == min_length:
            return valor

        printSubTitle(f"Valor inválido! Informe exatamente {min_length} caracteres.", "red")


def validate_input_str(mensagem):
    """
    Valida se o input tem o comprimento mínimo especificado.

    Args:
        mensagem: Mensagem para exibir ao usuário
        min_length: Comprimento mínimo exigido (padrão: 3)

    Returns:
        str: Input validado com comprimento mínimo
    """
    while True:
        try:
            valor = perguntar(mensagem, titulo="Entrada", obrigatorio=True, strip=True)
            if len(valor) < 2:
                printSubTitle("Erro, valor muito pequeno")
                continue
        except Exception:
            printSubTitle("Erro")
            continue
        return valor
            
def printTitle(
    message: str = "",
    color: str = "WHITE",
    screenSize: int = SCREEN_SIZE,
):
    colorama.reinit()
    color = COLORS[color]
    textToDisplay = str().join(list("-" for i in range(screenSize)))
    print("\n" + color + textToDisplay)

    auxSize = (screenSize - len(message)) / 2.0
    auxSize = auxSize if auxSize >= 0 else 0

    textToDisplay = str().join(list("-" for i in range(math.floor(auxSize))))
    textToDisplay += message
    textToDisplay += str().join(list("-" for i in range(math.ceil(auxSize))))
    print(color + textToDisplay)

    textToDisplay = str().join(list("-" for i in range(screenSize)))
    print(color + textToDisplay)

def debug(input_text: str, color: str = None, log: bool = True) -> None:
    """Print an input text in a specific color

    Args:
        color (str): Text color, according to the colors defined in termcolor lib

    """

    colorama.reinit()
    if log:
        print(colored(input_text, color)) if color is not None else print(input_text)

        


def salvar_dados_excel(dict_data: dict, nome_arquivo: str = "dados_manutencao.xlsx"):
    """
    Salva os dados do dicionário em uma planilha Excel.
    
    Args:
        dict_data: Dicionário com os dados a serem salvos
        nome_arquivo: Nome do arquivo Excel (padrão: "dados_manutencao.xlsx")
    """
    # Converter valores para string
    data_copy = copy.deepcopy(dict_data)
    for key, value in data_copy.items():
        data_copy[key] = str(value) if value is not None else "None"
    
    # Verificar se o arquivo já existe
    if os.path.exists(nome_arquivo):
        # Carregar workbook existente
        wb = load_workbook(nome_arquivo)
        ws = wb.active
        # Encontrar a próxima linha vazia
        next_row = ws.max_row + 1
    else:
        # Criar novo workbook
        wb = Workbook()
        ws = wb.active
        # Adicionar cabeçalho com as chaves do dicionário
        headers = list(data_copy.keys())
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
        next_row = 2
    
    # Adicionar dados na próxima linha
    for col_num, (key, value) in enumerate(data_copy.items(), 1):
        ws.cell(row=next_row, column=col_num, value=value)
    
    # Salvar o arquivo
    wb.save(nome_arquivo)
    printSubTitle(f"Dados salvos em {nome_arquivo}!", "green")

    
console = Console()


def format_value(value):
    text = str(value).lower()

    # ❌ erros
    if value is False or text in ["erro", "error", "fail", "falha"]:
        return "[bold red]❌ " + str(value) + "[/bold red]"

    # ✅ sucesso
    if value is True or text in ["ok", "true", "sucesso"]:
        return "[bold green]✅ " + str(value) + "[/bold green]"

    # ⚠️ aviso (opcional)
    if text in ["warning", "aviso", "atenção"]:
        return "[bold yellow]⚠ " + str(value) + "[/bold yellow]"

    # 🟡 padrão
    return "[yellow]" + str(value) + "[/yellow]"



def print_info(data: dict, title: str):
    table = Table(title=f"[bold yellow]{title}[/bold yellow]")

    table.add_column("Chave", style="cyan", no_wrap=True)
    table.add_column("Valor", justify="right")

    for k, v in data.items():
        value = v if v is not None else "None"

        # 🔴 se for erro → linha inteira vermelha
        if v is False:
            table.add_row(
                str(k),
                format_value(value),
                style="bold red"
            )
        else:
            table.add_row(
                str(k),
                format_value(value),
            )

    console.print(table)

def time_stamp():
    return datetime.now().strftime("%d/%m/%Y")



def guia_manutencao_DUO(dyid, modelo):  
    while True:
        try:
            printSubTitle(f" Guia de Manutenção Gateway {modelo} ", "cyan")
            dyid = dyid
            modelo = modelo
            versao_placa = validate_input_min_length("Digite a Versão da PCB Ex:100, 200, 210: ")
            versao_placa_completa = f"v{versao_placa[0]}.{versao_placa[1]}.{versao_placa[2]}"
            print(colored(f"Versão da PCB: {versao_placa_completa}", "cyan"))
            FIX_atualizado = validate_info(colored("PCB com FIX atualizado?", "cyan"))
            componentes_ok = validate_info(colored("Todos os componentes estão OK?", "cyan"))
            if componentes_ok == False:
                componentes_avariados = perguntar(colored("Digite quais os componentes não estão OK Ex: R2,R5:", "cyan"), titulo="Entrada", obrigatorio=False, strip=True)
            else:
                componentes_avariados = "Não há componentes avariados"
            fontes_3v8_3v3 = validate_info(colored("As fontes 3.8V e 3.3V estão OK(Sem curto)?: ", "cyan"))
            SD_card = validate_info(colored("O SD card está presente?: ", "cyan"))
            placa_limpa = validate_info(colored("PCB livre de químicos e/ou resíduos?: ", "cyan"))
            if modelo == "DUO" or modelo == "DUO-EX":
                fonte_intelbras = validate_info(colored("Gateway ja está com a Fonte Intelbras?: ", "cyan"))
                if fonte_intelbras == False:
                    fonte_ac_solda = validate_info(colored("Fonte AC: solda e pinos ok?: ", "cyan")) 
                else:
                    fonte_ac_solda = "Fonte Intelbras instalada"
                led_power = validate_info(colored("Led de Power (vermelho) está ligado?: ", "cyan"))
                fonte_220V = validate_info(colored("Fonte 220V: saída 12V ok?: ", "cyan"))
            else:
                fonte_220V = None
                fonte_intelbras = None
                fonte_ac_solda = None
                led_power = None
            rtc_bateria = validate_info(colored("bateria do RTC presente e bateria com carga?: ", "cyan"))
            tensao_24vdc = validate_info(colored("Alimentando com 24 Vdc, as tensões 3v8 e 3v3 estão OK?: ", "cyan"))
            consumo_24vdc = validate_info(colored("Alimentando com 24 Vdc, consumo OK?: ", "cyan"))
            pinos_u9 = validate_info(colored("Pinos 1 e 2 do CI U9 apresentam 1.8V?(1.8 so aparece se estiver configurado no interface MOBILE): ", "cyan"))
            led_wifi_ap = validate_info(colored("Led amarelo do WIFI - Modo AP OK?: ", "cyan"))
            acesso_ap = validate_info(colored("Foi possível acessar o modo AP?: ", "cyan"))
            imei = perguntar(colored("Digite o IMEI do dispositivo ou leia o Qr Code do TELIT: ", "cyan"), titulo="Entrada", obrigatorio=False, strip=True)
            if acesso_ap == True:
                comandos_at = validate_info(colored("Comandos AT do telit OK?: ", "cyan"))
                sinal_telit = validate_info(colored("A intensidade do sinal do Telit está ok?: ", "cyan"))
                iccid_simcard = validate_info(colored("Telit leu ICCID do SIMCARD?: ", "cyan"))
                echo_telit = validate_info(colored("O ECHO do Telit está desligado?: ", "cyan"))
                versao_firmware_main = validate_input_min_length("Digite a Versão de firmware DYG Main - STM: EX: 610 ")
                versao_firmware_main = f"v0{versao_firmware_main[0]}.0{versao_firmware_main[1]}.0{versao_firmware_main[2]}"
                print(colored(f"Versão de firmware DYG Main - STM: {versao_firmware_main}", "cyan"))
                versao_firmware_comm = validate_input_min_length("\nDigite a Versão de firmware DYG COMM - ESP: EX: 210 ")
                versao_firmware_comm = f"v0{versao_firmware_comm[0]}.0{versao_firmware_comm[1]}.0{versao_firmware_comm[2]}"
                print(colored(f"Versão de firmware DYG COMM - ESP: {versao_firmware_comm}", "cyan"))
            else:
                versao_firmware_comm = "Não foi possivel acessar o Modo AP"
                versao_firmware_main = "Não foi possivel acessar o Modo AP"
                echo_telit = "Não foi possivel acessar o Modo AP"
                sinal_telit = "Não foi possivel acessar o Modo AP"
                iccid_simcard = "Não foi possivel acessar o Modo AP"
                acesso_ap = "Não foi possivel acessar o Modo AP" 
                comandos_at = "Não foi possivel acessar o Modo AP"
            defeito_encontrado = perguntar("Qual o defeito encontrado?: ", titulo="Entrada", obrigatorio=False, strip=True)
            acao_realizada = perguntar("Qual ação realizada para resolver o defeito?: ", titulo="Entrada", obrigatorio=False, strip=True)
            atualizacoes = validate_info("Houve atualizações?: ")
            if atualizacoes:
                lista_atualizacoes = ["FIX", "Firmware", "Fonte 220V", "SDcard"]
                print(colored("Selecione as atualizações realizadas (digite os números separados por vírgula ou espaço):", "cyan"))
                for i, opcao in enumerate(lista_atualizacoes, 1):
                    print(colored(f"[{i}] - {opcao}", "cyan"))
                
                while True:
                    selecao = perguntar("Digite os números das opções Ex: 1 2 3: ", titulo="Entrada", obrigatorio=False, strip=True)
                    # Converter entrada para lista de números
                    numeros = []
                    for item in selecao.replace(',', ' ').split():
                        if item.strip().isdigit():
                            num = int(item.strip())
                            if 1 <= num <= len(lista_atualizacoes):
                                numeros.append(num)
                    
                    if numeros:
                        selecionados = [lista_atualizacoes[num - 1] for num in numeros]
                        atualizacoes = ", ".join(selecionados)
                        break
                    else:
                        printSubTitle("Seleção inválida! Tente novamente.", "red")
            else:
                atualizacoes = "Placa ja com os componentes atualizados"
        except KeyboardInterrupt:
            printSubTitle("Operação cancelada pelo usuário", "red")
            continue

        dados_finais = {
            "dyid": dyid,
            "modelo": modelo,
            "versao_placa": versao_placa_completa,
            "fix_atualizado": FIX_atualizado,
            "componentes_ok": componentes_ok,
            "componentes_avariados": componentes_avariados,
            "fontes_3v8_3v3": fontes_3v8_3v3,
            "sd_card": SD_card,
            "placa_limpa": placa_limpa,
            "fonte_intelbras": fonte_intelbras,
            "fonte_ac_solda": fonte_ac_solda,
            "led_power": led_power,
            "fonte_220V": fonte_220V,
            "tensao_24vdc": tensao_24vdc,
            "rtc_bateria": rtc_bateria,
            "consumo_24vdc": consumo_24vdc,
            "pinos_u9": pinos_u9,
            "led_wifi_ap": led_wifi_ap,
            "acesso_ap": acesso_ap,
            "imei": imei,
            "comandos_at": comandos_at,
            "sinal_telit": sinal_telit,
            "iccid_simcard": iccid_simcard,
            "echo_telit": echo_telit,
            "versao_firmware_main": versao_firmware_main,
            "versao_firmware_comm": versao_firmware_comm,
            "defeito_encontrado": defeito_encontrado,
            "acao_realizada": acao_realizada,
            "atualizacoes": atualizacoes,
            "timestamp": time_stamp(),
            "manutenção": None,
            "testes": None
        }           
        return dados_finais


def guia_manutencao_V2(dyid, modelo):  
    
        while True:
            try:
                printSubTitle(f" Guia de Manutenção Gateway {modelo} ", "cyan")
                dyid = dyid
                modelo = modelo
                versao_placa = validate_input_min_length("Digite a Versão da PCB Ex:100, 200, 210: ")
                versao_placa_completa = f"v{versao_placa[0]}.{versao_placa[1]}.{versao_placa[2]}"
                print(colored(f"Versão da PCB: {versao_placa_completa}", "cyan"))
                FIX_atualizado = None
                componentes_ok = validate_info(colored("Todos os componentes estão OK?", "cyan"))
                if componentes_ok == False:
                    componentes_avariados = perguntar(colored("Digite quais os componentes não estão OK Ex: R2,R5:", "cyan"), titulo="Entrada", obrigatorio=False, strip=True)
                else:
                    componentes_avariados = "Não há componentes avariados"
                if modelo == "IOT" or modelo == "LTE":
                    fonte_220V = validate_info(colored("Fonte 220V: saída 12V ok?: ", "cyan"))
                else:
                    fonte_220V = None
                SD_card = validate_info(colored("O SD card está presente?: ", "cyan"))
                placa_limpa = validate_info(colored("PCB livre de químicos e/ou resíduos?: ", "cyan"))
                fonte_intelbras = None
                fonte_ac_solda = None
                fontes_3v8_3v3 = validate_info(colored("As fontes 3.8V e 3.3V estão OK(Sem curto)?: ", "cyan"))
                led_power = validate_info(colored("Led de Power (vermelho) está ligado?: ", "cyan"))
                rtc_bateria = validate_info(colored("bateria do RTC presente e bateria com carga?: ", "cyan"))
                tensao_24vdc = validate_info(colored("Alimentando com 24 Vdc, as tensões 3v8 e 3v3 estão OK?: ", "cyan"))
                consumo_24vdc = validate_info(colored("Alimentando com 24 Vdc, consumo OK?: ", "cyan"))
                pinos_u9 = None
                led_wifi_ap = validate_info(colored("Led amarelo do WIFI - Modo AP OK?: ", "cyan"))
                acesso_ap = validate_info(colored("Foi possível acessar o modo AP?: ", "cyan"))
                imei = perguntar(colored("Digite o IMEI do dispositivo ou leia o Qr Code do TELIT: ", "cyan"), titulo="Entrada", obrigatorio=False, strip=True)
                if acesso_ap == True:
                    comandos_at = None
                    sinal_telit = None
                    iccid_simcard = validate_info(colored("Telit leu ICCID do SIMCARD?: ", "cyan"))
                    echo_telit = None
                    versao_firmware_main = validate_input_min_length("Digite a Versão de firmware DYG Main - STM: EX: 610 ")
                    versao_firmware_main = f"v0{versao_firmware_main[0]}.0{versao_firmware_main[1]}.0{versao_firmware_main[2]}"
                    print(colored(f"Versão de firmware DYG Main - STM: {versao_firmware_main}", "cyan"))
                    versao_firmware_comm = validate_input_min_length("\nDigite a Versão de firmware DYG COMM - ESP: EX: 210 ")
                    versao_firmware_comm = f"v0{versao_firmware_comm[0]}.0{versao_firmware_comm[1]}.0{versao_firmware_comm[2]}"
                    print(colored(f"Versão de firmware DYG COMM - ESP: {versao_firmware_comm}", "cyan"))
                else:
                    versao_firmware_comm = "Não foi possivel acessar o Modo AP"
                    versao_firmware_main = "Não foi possivel acessar o Modo AP"
                    iccid_simcard = "Não foi possivel acessar o Modo AP"
                    acesso_ap = "Não foi possivel acessar o Modo AP" 
                defeito_encontrado = perguntar("Qual o defeito encontrado?: ", titulo="Entrada", obrigatorio=False, strip=True)
                acao_realizada = perguntar("Qual ação realizada para resolver o defeito?: ", titulo="Entrada", obrigatorio=False, strip=True)
                atualizacoes = validate_info("Houve atualizações?: ")
                if atualizacoes:
                    lista_atualizacoes = ["Firmware"]
                    print(colored("Selecione as atualizações realizadas (digite os números separados por vírgula ou espaço):", "cyan"))
                    for i, opcao in enumerate(lista_atualizacoes, 1):
                        print(colored(f"[{i}] - {opcao}", "cyan"))
                    
                    while True:
                        selecao = perguntar("Digite os números das opções Ex: 1 2 3: ", titulo="Entrada", obrigatorio=False, strip=True)
                        # Converter entrada para lista de números
                        numeros = []
                        for item in selecao.replace(',', ' ').split():
                            if item.strip().isdigit():
                                num = int(item.strip())
                                if 1 <= num <= len(lista_atualizacoes):
                                    numeros.append(num)
                        
                        if numeros:
                            selecionados = [lista_atualizacoes[num - 1] for num in numeros]
                            atualizacoes = ", ".join(selecionados)
                            break
                        else:
                            printSubTitle("Seleção inválida! Tente novamente.", "red")
                else:
                    atualizacoes = "Placa ja com os componentes atualizados"
            except KeyboardInterrupt:
                printSubTitle("Operação cancelada pelo usuário", "red")
                continue


            dados_finais = {
            "dyid": dyid,
            "modelo": modelo,
            "versao_placa": versao_placa_completa,
            "fix_atualizado": FIX_atualizado,
            "componentes_ok": componentes_ok,
            "componentes_avariados": componentes_avariados,
            "fontes_3v8_3v3": fontes_3v8_3v3,
            "sd_card": SD_card,
            "placa_limpa": placa_limpa,
            "fonte_intelbras": fonte_intelbras,
            "fonte_ac_solda": fonte_ac_solda,
            "led_power": led_power,
            "fonte_220V": fonte_220V,
            "tensao_24vdc": tensao_24vdc,
            "rtc_bateria": rtc_bateria,
            "consumo_24vdc": consumo_24vdc,
            "pinos_u9": pinos_u9,
            "led_wifi_ap": led_wifi_ap,
            "acesso_ap": acesso_ap,
            "imei": imei,
            "comandos_at": comandos_at,
            "sinal_telit": sinal_telit,
            "iccid_simcard": iccid_simcard,
            "echo_telit": echo_telit,
            "versao_firmware_main": versao_firmware_main,
            "versao_firmware_comm": versao_firmware_comm,
            "defeito_encontrado": defeito_encontrado,
            "acao_realizada": acao_realizada,
            "atualizacoes": atualizacoes,
            "timestamp": time_stamp(),
            "manutenção": None,
            "testes": None
        }           
            return dados_finais
        
def guia_manutencao_EWI(dyid, modelo):  
    while True:
        try:
            printSubTitle(f" Guia de Manutenção Gateway {modelo} ", "cyan")
            dyid = dyid
            modelo = modelo
            versao_placa = validate_input_min_length("Digite a Versão da PCB Ex:100, 200, 210: ")
            versao_placa_completa = f"v{versao_placa[0]}.{versao_placa[1]}.{versao_placa[2]}"
            print(colored(f"Versão da PCB: {versao_placa_completa}", "cyan"))
            FIX_atualizado = validate_info(colored("PCB com FIX atualizado?", "cyan"))
            componentes_ok = validate_info(colored("Todos os componentes estão OK?", "cyan"))
            if componentes_ok == False:
                componentes_avariados = perguntar(colored("Digite quais os componentes não estão OK Ex: R2,R5:", "cyan"), titulo="Entrada", obrigatorio=False, strip=True)
            else:
                componentes_avariados = "Não há componentes avariados"
            fontes_3v8_3v3 = validate_info(colored("As fonte 3.3V esta OK(Sem curto)?: ", "cyan"))
            SD_card = validate_info(colored("O SD card está presente?: ", "cyan"))
            placa_limpa = validate_info(colored("PCB livre de químicos e/ou resíduos?: ", "cyan"))
            if modelo == "EWI":
                fonte_220V = validate_info(colored("Fonte 220V: saída 12V ok?: ", "cyan"))
                fonte_intelbras = validate_info(colored("Gateway ja está com a Fonte Intelbras?: ", "cyan"))
                if fonte_intelbras == False:
                    fonte_ac_solda = validate_info(colored("Fonte AC: solda e pinos ok?: ", "cyan")) 
                else:
                    fonte_ac_solda = "Fonte Intelbras instalada"
            else:
                fonte_220V = None
                fonte_intelbras = None
                fonte_ac_solda = None
                led_power = None
            rtc_bateria = validate_info(colored("bateria do RTC presente e bateria com carga?: ", "cyan"))
            tensao_24vdc = validate_info(colored("Alimentando com 24 Vdc, as tensões 3v8 e 3v3 estão OK?: ", "cyan"))
            consumo_24vdc = validate_info(colored("Alimentando com 24 Vdc, consumo OK?: ", "cyan"))
            pinos_u9 = None
            led_wifi_ap = validate_info(colored("Led amarelo do WIFI - Modo AP OK?: ", "cyan"))
            acesso_ap = validate_info(colored("Foi possível acessar o modo AP?: ", "cyan"))
            imei = None
            if acesso_ap == True:
                comandos_at = None
                sinal_telit = None
                iccid_simcard = None
                echo_telit = None
                versao_firmware_main = validate_input_min_length("Digite a Versão de firmware DYG Main - STM: EX: 610 ")
                versao_firmware_main = f"v0{versao_firmware_main[0]}.0{versao_firmware_main[1]}.0{versao_firmware_main[2]}"
                print(colored(f"Versão de firmware DYG Main - STM: {versao_firmware_main}", "cyan"))
                versao_firmware_comm = validate_input_min_length("\nDigite a Versão de firmware DYG COMM - ESP: EX: 210 ")
                versao_firmware_comm = f"v0{versao_firmware_comm[0]}.0{versao_firmware_comm[1]}.0{versao_firmware_comm[2]}"
                print(colored(f"Versão de firmware DYG COMM - ESP: {versao_firmware_comm}", "cyan"))
            else:
                versao_firmware_comm = "Não foi possivel acessar o Modo AP"
                versao_firmware_main = "Não foi possivel acessar o Modo AP"
                echo_telit = None
                sinal_telit = None
                iccid_simcard = None
                acesso_ap = "Não foi possivel acessar o Modo AP" 
                comandos_at = None
            defeito_encontrado = perguntar("Qual o defeito encontrado?: ", titulo="Entrada", obrigatorio=False, strip=True)
            acao_realizada = perguntar("Qual ação realizada para resolver o defeito?: ", titulo="Entrada", obrigatorio=False, strip=True)
            atualizacoes = validate_info("Houve atualizações?: ")
            if atualizacoes:
                lista_atualizacoes = ["FIX", "Firmware", "Fonte 220V", "SDcard"]
                print(colored("Selecione as atualizações realizadas (digite os números separados por vírgula ou espaço):", "cyan"))
                for i, opcao in enumerate(lista_atualizacoes, 1):
                    print(colored(f"[{i}] - {opcao}", "cyan"))
                
                while True:
                    selecao = perguntar("Digite os números das opções Ex: 1 2 3: ", titulo="Entrada", obrigatorio=False, strip=True)
                    # Converter entrada para lista de números
                    numeros = []
                    for item in selecao.replace(',', ' ').split():
                        if item.strip().isdigit():
                            num = int(item.strip())
                            if 1 <= num <= len(lista_atualizacoes):
                                numeros.append(num)
                    
                    if numeros:
                        selecionados = [lista_atualizacoes[num - 1] for num in numeros]
                        atualizacoes = ", ".join(selecionados)
                        break
                    else:
                        printSubTitle("Seleção inválida! Tente novamente.", "red")
            else:
                atualizacoes = "Placa ja com os componentes atualizados"
        except KeyboardInterrupt:
            printSubTitle("Operação cancelada pelo usuário", "red")
            continue
        dados_finais = {
                "dyid": dyid,
                "modelo": modelo,
                "versao_placa": versao_placa_completa,
                "fix_atualizado": FIX_atualizado,
                "componentes_ok": componentes_ok,
                "componentes_avariados": componentes_avariados,
                "fontes_3v8_3v3": fontes_3v8_3v3,
                "sd_card": SD_card,
                "placa_limpa": placa_limpa,
                "fonte_intelbras": fonte_intelbras,
                "fonte_ac_solda": fonte_ac_solda,
                "led_power": led_power,
                "fonte_220V": fonte_220V,
                "tensao_24vdc": tensao_24vdc,
                "rtc_bateria": rtc_bateria,
                "consumo_24vdc": consumo_24vdc,
                "pinos_u9": pinos_u9,
                "led_wifi_ap": led_wifi_ap,
                "acesso_ap": acesso_ap,
                "imei": imei,
                "comandos_at": comandos_at,
                "sinal_telit": sinal_telit,
                "iccid_simcard": iccid_simcard,
                "echo_telit": echo_telit,
                "versao_firmware_main": versao_firmware_main,
                "versao_firmware_comm": versao_firmware_comm,
                "defeito_encontrado": defeito_encontrado,
                "acao_realizada": acao_realizada,
                "atualizacoes": atualizacoes,
                "timestamp": time_stamp(),
                "manutenção": None,
                "testes": None
            }           
        return dados_finais
def somar_dias_uteis(qtd_dias):
    data = datetime.now()
    dias_adicionados = 0
    while dias_adicionados < qtd_dias:

        data += timedelta(days=1)
        if data.weekday() < 5:
            dias_adicionados += 1



    return data.strftime("%d/%m")
    

def entrada_de_garntia():
    lista_de_dados = []
    while True:
        dyid = validate_input_str("Digite o DYID: ")
        if not validate_dyid(dyid):
            continue
        printSubTitle(f"{dyid} Valido", "green")
        cliente = validate_input_str("Digite o cliente")
        os = validate_input_min_length("Digite o Numero da Ordem de serviço(OS): ", min_length=4)
        


        dados = {
            "dyid" : dyid,
            "cliente" : cliente,
            "ordem_de_servico" : os,
            # "data_de_entrada" : time_stamp(),
            # "prazo_de_conclusao" : somar_dias_uteis(15),
        }
        print_info(dados, "Dados Garantia")
        lista_de_dados.append(dados)
        if validate_info("Deseja adicionar mais gateways: "):
            continue
        else:
            return lista_de_dados



def enviar_card_teams(garantias):
    WEBHOOK_URL = r"https://default46837cc50aca4e61a3176b2d1f0bef.c1.environment.api.powerplatform.com:443/powerautomate/automations/direct/cu/08/workflows/5cb0ed05d6b84decb7499dce76844ab8/triggers/manual/paths/invoke?api-version=1&sp=%2Ftriggers%2Fmanual%2Frun&sv=1.0&sig=1k1_qpLmgptvMPx0qrtIRj_mgrjeVttIEYAttSv_tz4"

    rows = []

    # Cabeçalho
    rows.append({
        "type": "TableRow",
        "cells": [
            {
                "type": "TableCell",
                "items": [
                    {"type": "TextBlock", "text": "DYID", "weight": "Bolder"}
                ]
            },
            {
                "type": "TableCell",
                "items": [
                    {"type": "TextBlock", "text": "CLIENTE", "weight": "Bolder"}
                ]
            },
            {
                "type": "TableCell",
                "items": [
                    {"type": "TextBlock", "text": "OS", "weight": "Bolder"}
                ]
            },
            
        ]
    })

    # Dados
    for item in garantias:
        rows.append({
            "type": "TableRow",
            "cells": [
                {
                    "type": "TableCell",
                    "items": [
                        {"type": "TextBlock", "text": str(item["dyid"])}
                    ]
                },
                {
                    "type": "TableCell",
                    "items": [
                        {"type": "TextBlock", "text": str(item["cliente"])}
                    ]
                },
                {
                    "type": "TableCell",
                    "items": [
                        {"type": "TextBlock", "text": str(item["ordem_de_servico"])}
                    ]
                },
               
                
            ]
        })

    payload = {
        "type": "AdaptiveCard",
        "$schema": "http://adaptivecards.io/schemas/adaptive-card.json",
        "version": "1.4",
        "body": [
            {
                "type": "TextBlock",
                "text": "🚨 Entradas de Garantia",
                "weight": "Bolder",
                "size": "Large"
            },
            {
                "type": "Table",
                "showGridLines": True,
                "firstRowAsHeaders": True,
                "columns": [
                    {"width": 5},  # DYIDS
                    {"width": 10},  # CLIENTE
                    {"width": 2},  # OS
                    # {"width": 2},  # ENTRADA
                    # {"width": 2}   # ENTREGA
                ],
                "rows": rows
            }
        ]
    }

    response = requests.post(
        WEBHOOK_URL,
        json=payload
    )

    if response.status_code == 202:
        printSubTitle(" Mensagem enviada com SUCESSO ", "green")
    else:
        printSubTitle(" Error, algo deu errado ao envia a mensagem ", "red")



console = Console()

if PROMPT_TOOLKIT_AVAILABLE:
    style = Style.from_dict({
        "prompt": "#00d7ff bold",
        "": "#ffffff",
    })

    history = InMemoryHistory()
else:
    style = None
    history = None


def perguntar(pergunta: str,
              titulo: str = "📋 Sistema",
              obrigatorio: bool = True,
              upper: bool = False,
              strip: bool = True):

    if PROMPT_TOOLKIT_AVAILABLE:
        class Validador(Validator):
            def validate(self, document):
                texto = document.text

                if obrigatorio and not texto.strip():
                    raise ValidationError(
                        message="Este campo é obrigatório.",
                        cursor_position=0
                    )

        console.print()

        console.print(
            Panel(
                Align.left(f"[bold cyan]{pergunta}[/bold cyan]"),
                title=f"[bold green]{titulo}[/bold green]",
                border_style="bright_blue",
                padding=(1, 2),
            )
        )

        resposta = prompt(
            [("class:prompt", "❯ Resposta: ")],
            validator=Validador(),
            validate_while_typing=False,
            history=history,
            style=style
        )
    else:
        console.print()
        console.print(
            Panel(
                Align.left(f"[bold cyan]{pergunta}[/bold cyan]"),
                title=f"[bold green]{titulo}[/bold green]",
                border_style="bright_blue",
                padding=(1, 2),
            )
        )
        resposta = perguntar(pergunta, titulo=titulo, obrigatorio=obrigatorio, upper=upper, strip=strip)

    if strip:
        resposta = resposta.strip()

    if upper:
        resposta = resposta.upper()

    console.print()

    return resposta


